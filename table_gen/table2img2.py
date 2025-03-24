import os
import random
import string
import pandas as pd
import dataframe_image as dfi
from PIL import Image, ImageDraw, ImageFont, ImageFile
import subprocess
import tempfile
import math
from openpyxl import load_workbook
from openpyxl.styles.colors import Color  # Import Color class
import json
import shutil

table_error_count = 0
ImageFile.LOAD_TRUNCATED_IMAGES = True
Image.MAX_IMAGE_PIXELS = None


def generate_table_image(table_data, table_name, output_dir):
    """
    Generates an image from a JSON table with the table name overlaid using pandas and dataframe_image.
    Randomly selects a table style from a diverse set of styles.
    Handles large tables by splitting them into multiple images and combining them horizontally.

    Args:
        table_data: JSON data representing the table.
        table_name: The name of the table.
        output_dir: Directory to save the generated image.
    Returns:
        A tuple containing the image ID and the path to the saved image.
    """
    global table_error_count

    df = None
    if "table_columns" in table_data and "table_content" in table_data:
        df = pd.DataFrame(
            table_data["table_content"], columns=table_data["table_columns"]
        )
    else:
        df = pd.DataFrame(table_data)

    num_rows, num_cols = df.shape

    # Generate a 5-character random image ID
    image_id = "".join(
        random.choices(string.ascii_lowercase + string.digits, k=5)
    ).capitalize()
    image_filename = f"TableImg_{image_id}_{num_rows}.png"
    image_path = os.path.join(output_dir, image_filename)

    # Define a function for alternating row colors
    def alternating_row_colors(df, color1="#f2f2f2", color2="white"):
        return [
            f"background-color: {color1}" if i % 2 else f"background-color: {color2}"
            for i in range(len(df))
        ]

    # Define a variety of table styles
    table_styles = [
        # Style 1: Green header with bold font, alternating row colors
        lambda df: df.style.apply(alternating_row_colors, axis=0).set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#4CAF50"),
                        ("color", "white"),
                        ("font-weight", "bold"),
                        ("text-align", "center"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [("border", "1px solid black"), ("text-align", "center")],
                },
                {"selector": "table", "props": [("border-collapse", "collapse")]},
            ]
        ),
        # Style 2: Minimalist style with light grey background
        lambda df: df.style.set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#E0E0E0"),
                        ("color", "black"),
                        ("text-align", "center"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [("text-align", "center"), ("background-color", "white")],
                },
            ]
        ),
        # Style 3: Dark header with white text, bold rows
        lambda df: df.style.apply(
            alternating_row_colors, axis=0, color1="#E6E6E6", color2="white"
        ).set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#333"),
                        ("color", "white"),
                        ("text-align", "center"),
                        ("border", "1px solid white"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [("border", "1px solid #999"), ("text-align", "center")],
                },
            ]
        ),
        # Style 4: Blue header with bold white text and gridlines
        lambda df: df.style.apply(
            alternating_row_colors, axis=0, color1="#F8F8F8", color2="#FFFFFF"
        ).set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#1E88E5"),
                        ("color", "white"),
                        ("font-weight", "bold"),
                        ("text-align", "center"),
                        ("border", "2px solid #000"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [("border", "1px solid #888"), ("text-align", "center")],
                },
            ]
        ),
        # Style 5: Modern white and grey with no borders
        lambda df: df.style.set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#F0F0F0"),
                        ("font-size", "12pt"),
                        ("text-align", "center"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [
                        ("background-color", "#FAFAFA"),
                        ("text-align", "center"),
                    ],
                },
            ]
        ),
        # Style 6: Orange header with rounded borders
        lambda df: df.style.set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#FF7043"),
                        ("color", "white"),
                        ("border", "1px solid #FF5722"),
                        ("text-align", "center"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [
                        ("text-align", "center"),
                        ("border", "1px solid #FFCCBC"),
                    ],
                },
            ]
        ),
        # Style 7: Red header with bold font, striped rows
        lambda df: df.style.apply(
            alternating_row_colors, axis=0, color1="#FFEBEE", color2="#FFFFFF"
        ).set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#B71C1C"),
                        ("color", "white"),
                        ("font-weight", "bold"),
                        ("text-align", "center"),
                    ],
                },
                {"selector": "td", "props": [("text-align", "center")]},
            ]
        ),
        # Style 8: Elegant black-and-white style
        lambda df: df.style.set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "black"),
                        ("color", "white"),
                        ("text-align", "center"),
                        ("border", "1px solid white"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [
                        ("background-color", "white"),
                        ("color", "black"),
                        ("border", "1px solid black"),
                        ("text-align", "center"),
                    ],
                },
            ]
        ),
        # Style 9: Pastel blue theme with centered text
        lambda df: df.style.set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#BBDEFB"),
                        ("color", "#0D47A1"),
                        ("text-align", "center"),
                        ("font-weight", "bold"),
                    ],
                },
                {"selector": "td", "props": [("text-align", "center")]},
            ]
        ),
        # Style 10: Vibrant yellow header with alternating bold text rows
        lambda df: df.style.apply(
            alternating_row_colors, axis=0, color1="#FFFDE7", color2="#FFF9C4"
        ).set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#FBC02D"),
                        ("color", "black"),
                        ("font-weight", "bold"),
                        ("text-align", "center"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [
                        ("border", "1px solid #F57F17"),
                        ("text-align", "center"),
                    ],
                },
            ]
        ),
    ]

    # Randomly select a style
    selected_style = random.choice(table_styles)

    # Define possible conversions with custom probabilities
    conversion_options = [("selenium", 0.7), ("matplotlib", 0.3)]

    # Select one based on the given probabilities
    table_conversion = random.choices(
        [option[0] for option in conversion_options],
        weights=[option[1] for option in conversion_options],
        k=1,
    )[0]

    # Split large tables into multiple parts
    if num_rows > 80:
        df = df.head(80)
        num_rows = 80
        image_filename = f"TableImg_{image_id}_{num_rows}.png"
        image_path = os.path.join(output_dir, image_filename)

    if num_rows >= 80:
        part_images = []
        part_size = math.ceil(num_rows / math.ceil(num_rows / 30))
        for i in range(0, num_rows, part_size):
            df_part = df.iloc[i : i + part_size]
            styled_df_part = selected_style(df_part).hide(axis="index")

            part_image_filename = f"TableImg_{image_id}_{i}.png"
            part_image_path = os.path.join(output_dir, part_image_filename)

            with tempfile.TemporaryDirectory() as temp_dir:
                try:
                    temp_image_path = os.path.join(
                        temp_dir, f"temp_table_{image_id}_{i}.png"
                    )
                    dfi.export(
                        styled_df_part,
                        temp_image_path,
                        table_conversion=table_conversion,
                        max_rows=-1,
                        max_cols=-1,
                    )
                    subprocess.run(["cp", temp_image_path, part_image_path])
                    part_images.append(part_image_path)

                except Exception as e:
                    table_error_count += 1
                    print(
                        f"Error exporting table part to image: {e}, Total errors: {table_error_count}"
                    )
                    return None, None

        # Combine part images horizontally
        padding = 10  # Padding between images
        images = [Image.open(p) for p in part_images]
        total_width = sum(image.width for image in images) + padding * (len(images) - 1)
        max_height = max(image.height for image in images)

        combined_img = Image.new("RGB", (total_width, max_height), color="white")
        x_offset = 0
        for img in images:
            combined_img.paste(img, (x_offset, 0))
            x_offset += img.width + padding

        combined_img.save(image_path)

        # Clean up part images
        for p in part_images:
            os.remove(p)

    else:
        # Style the DataFrame (for smaller tables)
        styled_df = selected_style(df).hide(axis="index")

        # Use the library's default export for other conversions (html2image, matplotlib)
        with tempfile.TemporaryDirectory() as temp_dir:
            try:
                # Use a temporary file within the temporary directory
                temp_image_path = os.path.join(temp_dir, f"temp_table_{image_id}.png")

                # Use the library's default export for other conversions (html2image, matplotlib)
                dfi.export(
                    styled_df,
                    temp_image_path,
                    table_conversion=table_conversion,
                    max_rows=-1,
                    max_cols=-1,
                )
                subprocess.run(["cp", temp_image_path, image_path])
            except Exception as e:
                table_error_count += 1
                print(
                    f"Error exporting table to image: {e}, Total errors: {table_error_count}"
                )
                return None, None

    # Add table name above the image
    try:
        table_img = Image.open(image_path)
        table_width, table_height = table_img.size

        font = ImageFont.load_default()
        padding = 10
        table_name_upper = table_name.upper()

        draw = ImageDraw.Draw(table_img)
        text_bbox = draw.textbbox((0, 0), table_name_upper, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]

        new_img_width = table_width
        new_img_height = table_height + text_height + (2 * padding)
        new_img = Image.new("RGB", (new_img_width, new_img_height), "white")
        new_draw = ImageDraw.Draw(new_img)

        table_name_x = (table_width - text_width) // 2
        table_name_y = padding // 2

        new_draw.rectangle(
            [0, 0, new_img_width, text_height + 2 * padding], fill="white"
        )
        new_draw.text(
            (table_name_x, table_name_y), table_name_upper, font=font, fill="black"
        )

        new_img.paste(table_img, (0, text_height + 2 * padding))
        new_img.save(image_path)

    except Exception as e:
        print(f"Error adding table name to image: {e}")

    return image_filename, image_path


def get_pil_color(excel_color):
    """Convert Excel color to PIL color, handling different types and errors."""
    if excel_color is None:
        return "white"  # Default white if no fill

    if isinstance(excel_color, Color):  # Check if it's an openpyxl Color object
        if hasattr(excel_color, "rgb"):
            rgb_value = excel_color.rgb
            if isinstance(rgb_value, str):
                if len(rgb_value) == 8 and rgb_value.startswith(
                    "FF"
                ):  # Standard RGB hex format from openpyxl
                    return "#" + rgb_value[2:]  # Remove alpha prefix
                elif len(rgb_value) == 6:
                    return "#" + rgb_value
            elif isinstance(rgb_value, bytes):
                return f"#{rgb_value.hex()}"
        elif hasattr(excel_color, "index"):
            # Handle theme colors or indexed colors with a default
            return "lightgray"
        return "lightgray"

    elif isinstance(excel_color, tuple) and len(excel_color) == 3:
        r, g, b = excel_color
        return f"rgb({r},{g},{b})"

    elif isinstance(excel_color, str) and excel_color.startswith("#"):
        return excel_color

    else:
        print(
            f"Warning: Unexpected color type: {type(excel_color)}, value: {excel_color}. Using default gray."
        )
        return "lightgray"


def generate_sheet_image(sheet_df, sheet_name, output_dir, spreadsheet_filename):
    """
    Generates an image of an Excel sheet using openpyxl and Pillow, with dynamic cell sizing and basic styling.
    """
    global table_error_count
    image_id = "".join(
        random.choices(string.ascii_lowercase + string.digits, k=5)
    ).capitalize()

    # Truncate sheet_name to 31 characters if needed
    truncated_sheet_name = sheet_name[:31] if len(sheet_name) > 31 else sheet_name

    image_filename = f"SheetImg_{image_id}.png"
    image_path = os.path.join(output_dir, image_filename)

    try:
        # Load workbook using openpyxl
        workbook = load_workbook(spreadsheet_filename, data_only=True)
        worksheet = workbook[truncated_sheet_name]

        # --- Dynamic Cell Sizing and Styled Rendering with Pillow ---
        start_x, start_y = 20, 50  # Starting position
        padding = 5
        default_font = ImageFont.load_default()  # Still using default font for now
        col_widths = [0] * (worksheet.max_column + 1)  # List to store column widths

        # --- First Pass: Calculate Column Widths and Row Heights ---
        row_heights = [0] * (worksheet.max_row + 1)
        for row_idx, row in enumerate(
            worksheet.rows, start=1
        ):  # 1-based index from openpyxl
            max_text_height_in_row = 0
            for col_idx, cell in enumerate(row, start=1):  # 1-based index
                cell_value = str(cell.value) if cell.value is not None else ""
                text_bbox = ImageDraw.Draw(Image.new("RGB", (1, 1))).textbbox(
                    (0, 0), cell_value, font=default_font
                )  # Dummy image for bbox
                text_width = text_bbox[2] - text_bbox[0]
                text_height = text_bbox[3] - text_bbox[1]

                col_widths[col_idx] = max(
                    col_widths[col_idx], text_width + 2 * padding
                )  # Track max width for each column
                max_text_height_in_row = max(
                    max_text_height_in_row, text_height
                )  # Track max height in row

            row_heights[row_idx] = max_text_height_in_row + 2 * padding

        cell_width_CumSum = 0
        cumulative_col_widths = [0] * (worksheet.max_column + 1)
        for i in range(
            1, len(col_widths)
        ):  # Start from index 1 to align with column indices
            cumulative_col_widths[i] = cumulative_col_widths[i - 1] + col_widths[i]

        cumulative_row_heights = [0] * (worksheet.max_row + 1)
        for i in range(1, len(row_heights)):
            cumulative_row_heights[i] = cumulative_row_heights[i - 1] + row_heights[i]

        # Calculate image size based on calculated column widths and row heights
        image_width = start_x + cumulative_col_widths[worksheet.max_column] + start_x
        image_height = start_y + cumulative_row_heights[worksheet.max_row] + start_y
        img = Image.new("RGB", (image_width, image_height), "white")
        draw = ImageDraw.Draw(img)

        # --- Second Pass: Draw Cells with Styles ---
        for row_idx, row in enumerate(worksheet.rows, start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_value = str(cell.value) if cell.value is not None else ""

                x1 = (
                    start_x + cumulative_col_widths[col_idx - 1]
                )  # Use cumulative widths
                y1 = (
                    start_y + cumulative_row_heights[row_idx - 1]
                )  # Use cumulative heights
                x2 = x1 + col_widths[col_idx]
                y2 = y1 + row_heights[row_idx]

                # --- Apply Fill Color ---
                fill = cell.fill
                if fill and fill.fill_type == "solid":  # Check if there's a solid fill
                    fgColor = fill.fgColor
                    fill_color = get_pil_color(
                        fgColor
                    )  # Use the updated get_pil_color function
                    draw.rectangle(
                        [(x1, y1), (x2, y2)], fill=fill_color, outline="lightgray"
                    )  # Apply fill
                else:
                    draw.rectangle(
                        [(x1, y1), (x2, y2)], fill="white", outline="lightgray"
                    )  # Default white fill if no solid fill

                # --- Draw Cell Text ---
                text_x = x1 + padding
                text_y = y1 + padding
                draw.text(
                    (text_x, text_y), cell_value, fill="black", font=default_font
                )  # Basic black text

        img.save(image_path, "PNG")

    except Exception as e:
        table_error_count += 1
        print(
            f"OpenPyXL + Pillow Error (Styled): {e}, Spreadsheet: {spreadsheet_filename}, Sheet: {sheet_name}, Total errors: {table_error_count}"
        )
        return None, None

    # Add sheet name above the image (keeping this part)
    try:
        sheet_img = Image.open(image_path)
        sheet_width, sheet_height = sheet_img.size

        font = ImageFont.load_default()
        padding = 5
        sheet_name_upper = sheet_name.upper()

        draw = ImageDraw.Draw(sheet_img)
        text_bbox = draw.textbbox((0, 0), sheet_name_upper, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]

        new_img_width = sheet_width
        new_img_height = sheet_height + text_height + (2 * padding)
        new_img = Image.new("RGB", (new_img_width, new_img_height), "white")
        new_draw = ImageDraw.Draw(new_img)

        sheet_name_x = (sheet_width - text_width) // 2
        sheet_name_y = padding // 2

        new_draw.rectangle(
            [0, 0, new_img_width, text_height + 2 * padding], fill="white"
        )
        new_draw.text(
            (sheet_name_x, sheet_name_y), sheet_name_upper, font=font, fill="black"
        )

        new_img.paste(sheet_img, (0, text_height + 2 * padding))
        new_img.save(image_path)

    except Exception as e:
        print(
            f"PIL Error adding sheet name (Styled): {e}, Spreadsheet: {spreadsheet_filename}, Sheet: {sheet_name}"
        )

    return image_filename, image_path


def generate_table_image_multitab(table_data, table_name, output_dir):
    """
    Generates an image from table data for multi-tab datasets.
    Supports various input formats and ensures tables don't exceed the 50 row limit.

    Args:
        table_data: Table data in various possible formats (dict, DataFrame, string, etc.)
        table_name: The name of the table to display
        output_dir: Directory to save the generated image

    Returns:
        A tuple containing the image ID and the path to the saved image
    """
    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Check table_data type and convert to appropriate format
    df = None

    # Case 1: table_data is already a pandas DataFrame
    if isinstance(table_data, pd.DataFrame):
        df = table_data

    # Case 2: table_data is a dict with table_columns and table_content
    elif (
        isinstance(table_data, dict)
        and "table_columns" in table_data
        and "table_content" in table_data
    ):
        df = pd.DataFrame(
            table_data["table_content"], columns=table_data["table_columns"]
        )

    # Case 3: table_data is a dict with columns and data (common in ATIS)
    elif (
        isinstance(table_data, dict)
        and "columns" in table_data
        and "data" in table_data
    ):
        df = pd.DataFrame(table_data["data"], columns=table_data["columns"])

    # Case 4: table_data is a plain dict (convert keys to columns and values to rows)
    elif isinstance(table_data, dict):
        df = pd.DataFrame.from_dict(table_data)

    # Case 5: table_data is a string (try to parse as JSON)
    elif isinstance(table_data, str):
        try:
            # Try to parse as JSON
            parsed_data = json.loads(table_data)
            if isinstance(parsed_data, dict):
                if "table_columns" in parsed_data and "table_content" in parsed_data:
                    df = pd.DataFrame(
                        parsed_data["table_content"],
                        columns=parsed_data["table_columns"],
                    )
                elif "columns" in parsed_data and "data" in parsed_data:
                    df = pd.DataFrame(
                        parsed_data["data"], columns=parsed_data["columns"]
                    )
                else:
                    df = pd.DataFrame.from_dict(parsed_data)
            elif isinstance(parsed_data, list):
                df = pd.DataFrame(parsed_data)
        except json.JSONDecodeError:
            # If not valid JSON, try to parse as CSV
            try:
                df = pd.read_csv(pd.io.common.StringIO(table_data))
            except:
                print(f"Could not parse string table data: {table_data[:100]}...")
                return None, None

    # Case 6: table_data is a list (convert to DataFrame directly)
    elif isinstance(table_data, list):
        df = pd.DataFrame(table_data)

    # If we couldn't convert to DataFrame, return None
    if df is None:
        print(f"Could not convert table data to DataFrame. Type: {type(table_data)}")
        return None, None

    # Enforce 50 row limit on tables
    if len(df) > 50:
        df = df.head(50)

    # Clean table name for filename
    safe_table_name = "".join(c if c.isalnum() else "_" for c in table_name)
    safe_table_name = safe_table_name[:30]  # Limit filename length

    # Generate a unique image ID
    image_id = "".join(
        random.choices(string.ascii_lowercase + string.digits, k=5)
    ).capitalize()
    image_filename = f"TableImg_{image_id}_{safe_table_name}.png"
    image_path = os.path.join(output_dir, image_filename)

    # Define functions for table styling
    def alternating_row_colors(df, color1="#f2f2f2", color2="white"):
        return [
            f"background-color: {color1}" if i % 2 else f"background-color: {color2}"
            for i in range(len(df))
        ]

    # Define a variety of table styles
    table_styles = [
        # Style 1: Green header with bold font, alternating row colors
        lambda df: df.style.apply(alternating_row_colors, axis=0).set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#4CAF50"),
                        ("color", "white"),
                        ("font-weight", "bold"),
                        ("text-align", "center"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [("border", "1px solid black"), ("text-align", "center")],
                },
                {"selector": "table", "props": [("border-collapse", "collapse")]},
            ]
        ),
        # Style 2: Minimalist style with light grey background
        lambda df: df.style.set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#E0E0E0"),
                        ("color", "black"),
                        ("text-align", "center"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [("text-align", "center"), ("background-color", "white")],
                },
            ]
        ),
        # Style 3: Dark header with white text, bold rows
        lambda df: df.style.apply(
            alternating_row_colors, axis=0, color1="#E6E6E6", color2="white"
        ).set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#333"),
                        ("color", "white"),
                        ("text-align", "center"),
                        ("border", "1px solid white"),
                    ],
                },
                {
                    "selector": "td",
                    "props": [("border", "1px solid #999"), ("text-align", "center")],
                },
            ]
        ),
    ]

    # Randomly select a style
    selected_style = random.choice(table_styles)

    # Define possible conversions with custom probabilities
    conversion_options = [("selenium", 0.7), ("matplotlib", 0.3)]
    table_conversion = random.choices(
        [option[0] for option in conversion_options],
        weights=[option[1] for option in conversion_options],
        k=1,
    )[0]

    # Style the DataFrame and generate the image
    styled_df = selected_style(df).hide(axis="index")

    try:
        # Create temporary directory for image generation
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_image_path = os.path.join(temp_dir, f"temp_table_{image_id}.png")

            # Generate the table image
            dfi.export(
                styled_df,
                temp_image_path,
                table_conversion=table_conversion,
                max_rows=-1,
                max_cols=-1,
            )

            shutil.copy(temp_image_path, image_path)

            # Add table name overlay
            try:
                table_img = Image.open(image_path)
                table_width, table_height = table_img.size

                # Use default font since custom fonts might not be available
                font = ImageFont.load_default()
                padding = 10

                # Create a display name for the table (uppercase for visibility)
                table_name_display = table_name.upper()

                # Calculate text dimensions
                draw = ImageDraw.Draw(table_img)
                text_bbox = draw.textbbox((0, 0), table_name_display, font=font)
                text_width = text_bbox[2] - text_bbox[0]
                text_height = text_bbox[3] - text_bbox[1]

                # Create a new image with space for the title
                new_img_width = table_width
                new_img_height = table_height + text_height + (2 * padding)
                new_img = Image.new("RGB", (new_img_width, new_img_height), "white")
                new_draw = ImageDraw.Draw(new_img)

                # Calculate text position (centered)
                table_name_x = (table_width - text_width) // 2
                table_name_y = padding // 2

                # Draw title background and text
                new_draw.rectangle(
                    [0, 0, new_img_width, text_height + 2 * padding], fill="white"
                )
                new_draw.text(
                    (table_name_x, table_name_y),
                    table_name_display,
                    font=font,
                    fill="black",
                )

                # Paste the table image below the title
                new_img.paste(table_img, (0, text_height + 2 * padding))
                new_img.save(image_path)

            except Exception as e:
                print(f"Error adding table name to image: {e}")
                # If adding the name fails, just use the original image

            return image_filename, image_path

    except Exception as e:
        print(f"Error generating table image: {e}")
        return None, None
